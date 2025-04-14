import os
import logging
import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from dash import dcc
import io

# Configuração de logging detalhada
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Configuração de caminhos dinâmica
MOUNT_PATH = '/data' if os.environ.get('RENDER') else os.path.join(os.getcwd(), 'data')
EXCEL_PATH = os.path.join(MOUNT_PATH, 'b.xlsx')

def setup_persistent_environment():
    try:
        os.makedirs(MOUNT_PATH, exist_ok=True)

        if not os.path.exists(EXCEL_PATH):
            wb = Workbook()
            
            # Remove sheet padrão vazio se existir
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            
            # Cria aba JAN com cabeçalhos
            ws = wb.create_sheet("JAN")
            headers = [
                'data', 'beneficiario', 'valor_transacionado', 'valor_liberado',
                'taxa_de_juros', 'comissao_agente', 'extra_agente', 'valor_dualcred',
                'nota_fiscal', 'porcentagem_agente', 'quantidade_parcelas', 'agente',
                '%trans', '%liberad'
            ]
            ws.append(headers)
            
            # Cria outras abas mensais vazias
            months = ['FEV', 'MAR', 'ABR', 'MAI', 'JUN', 
                     'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']
            for month in months:
                ws = wb.create_sheet(month)
                ws.append(headers)
            
            wb.save(EXCEL_PATH)
        
        if not os.access(MOUNT_PATH, os.W_OK):
            logger.error(f"Sem permissão de escrita em: {MOUNT_PATH}")
            raise PermissionError("Erro de permissão no diretório persistente")

    except Exception as e:
        logger.error(f"Falha na configuração inicial: {str(e)}")
        raise

def sanitize_column_name(col):
    return (
        str(col)
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("ç", "c")
        .replace("ã", "a")
        .replace("õ", "o")
        .replace("ó", "o")
        .replace("ô", "o")
        .replace("à", "a")
        .replace("é", "e")
        .replace("ê", "e")
        .replace("ú", "u")
        .replace("%", "porcento")
        .replace("(", "")
        .replace(")", "")
    )

def load_and_process_data():
    """Carrega dados mantendo a estrutura por abas"""
    try:
        setup_persistent_environment()
        logger.info("Iniciando processamento de dados...")

        # Mapeamento de colunas
        column_mapping = {
            'beneficiário': 'beneficiario',
            'comissão_agente': 'comissao_agente',
            'chave_pix_cpf': 'chave_pix',
            '%_trans': '%trans',
            '%_liberad': '%liberad',
            'máquina': 'maquina'
        }

        # Carregar abas como dicionário de DataFrames
        sheets = pd.read_excel(EXCEL_PATH, sheet_name=None, engine='openpyxl')
        
        # Processar cada aba individualmente
        processed_sheets = {}
        for sheet_name, df in sheets.items():
            try:
                # Sanitizar e padronizar colunas
                df.columns = [sanitize_column_name(col) for col in df.columns]
                df.rename(columns=column_mapping, inplace=True, errors='ignore')
                
                # Adicionar colunas faltantes com valores padrão
                required_columns = [
                    'data', 'beneficiario', 'valor_transacionado', 'valor_liberado',
                    'taxa_de_juros', 'comissao_agente', 'extra_agente', 'valor_dualcred',
                    'nota_fiscal', 'porcentagem_agente', 'quantidade_parcelas', 'agente',
                    '%trans', '%liberad'
                ]
                
                for col in required_columns:
                    if col not in df.columns:
                        df[col] = pd.NaT if col == 'data' else 0.0
                # Cálculos condicionais
                df['valor_dualcred'] = (
                    df['valor_transacionado'] 
                    - df['valor_liberado'] 
                    - df['taxa_de_juros'] 
                    - df['comissao_agente'] 
                    - df['extra_agente']
                ).round(2)

                df['%trans'] = np.where(
                    df['valor_transacionado'] > 0,
                    (df['valor_dualcred'] / df['valor_transacionado']) * 100,
                    0
                ).round(2)

                df['%liberad'] = np.where(
                    df['valor_liberado'] > 0,
                    (df['valor_dualcred'] / df['valor_liberado']) * 100,
                    0
                ).round(2)

                df['nota_fiscal'] = (df['valor_transacionado'] * 0.032).round(2)

                # Ordenar colunas conforme layout original
                df = df.reindex(columns=[
                    'data', 'beneficiario', 'valor_transacionado', 'valor_liberado',
                    'taxa_de_juros', 'comissao_agente', 'extra_agente', 'valor_dualcred',
                    'nota_fiscal', 'porcentagem_agente', 'quantidade_parcelas', 'agente',
                    '%trans', '%liberad'
                ])

                processed_sheets[sheet_name] = df
                logger.info(f"Aba {sheet_name} processada com sucesso")

            except Exception as e:
                logger.error(f"Erro na aba {sheet_name}: {str(e)}")
                continue

        return processed_sheets  # Retorna dicionário de DataFrames

    except Exception as e:
        logger.error(f"Erro crítico: {str(e)}")
        return {}
    

def salvar_no_excel(df):
    """Salva o DataFrame dividindo as linhas por abas mensais."""
    try:
        logger.info("Salvando dados...")
        setup_persistent_environment()

        # Mapear meses para nomes das abas
        month_names = {
            1: 'JAN', 2: 'FEV', 3: 'MAR', 4: 'ABR', 5: 'MAI', 6: 'JUN',
            7: 'JUL', 8: 'AGO', 9: 'SET', 10: 'OUT', 11: 'NOV', 12: 'DEZ'
        }

        # Criar um writer para o Excel
        writer = pd.ExcelWriter(EXCEL_PATH, engine='openpyxl')

        # Dividir o DataFrame por mês e salvar em abas
        df['data'] = pd.to_datetime(df['data'])
        df['month'] = df['data'].dt.month.map(month_names)

        for sheet_name in month_names.values():
            # Filtrar dados do mês
            df_month = df[df['month'] == sheet_name].drop(columns=['month'])

            # Garantir a ordem das colunas
            df_month = df_month.reindex(columns=[
                'data', 'beneficiario', 'valor_transacionado', 'valor_liberado',
                'taxa_de_juros', 'comissao_agente', 'extra_agente', 'valor_dualcred',
                'nota_fiscal', 'porcentagem_agente', 'quantidade_parcelas', 'agente',
                '%trans', '%liberad'
            ])

            # Salvar na aba correspondente
            df_month.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

        writer.close()
        return True
    except Exception as e:
        logger.error(f"Erro ao salvar: {str(e)}")
        return False
    
def exportar_dados(processed_sheets):
    """Exporta mantendo a estrutura por abas"""
    try:
        logger.info("Iniciando exportação...")
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            for sheet_name, df in processed_sheets.items():
                logger.info(f"Exportando aba: {sheet_name}")
                
                # Verificar se df tem as colunas necessárias
                if df.empty:
                    logger.warning(f"Aba {sheet_name} vazia")
                    continue
                    
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    columns=[
                        'data', 'beneficiario', 'valor_transacionado', 'valor_liberado',
                        'taxa_de_juros', 'comissao_agente', 'extra_agente', 'valor_dualcred',
                        'nota_fiscal', 'quantidade_parcelas', 'agente', '%trans', '%liberad'
                    ]
                )
        
        buffer.seek(0)
        logger.info("Exportação concluída com sucesso")
        return dcc.send_bytes(
            buffer.getvalue(),
            filename="Dados_Atualizados.xlsx",
            type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    except Exception as e:
        logger.error(f"Erro na exportação: {str(e)}", exc_info=True)  # Log detalhado
        return None
    
# Inicialização segura
try:
    processed_sheets = load_and_process_data()
    if not processed_sheets:
        logger.warning("Nenhuma aba válida encontrada")
    else:
        logger.info(f"Dados carregados: {len(processed_sheets)} abas")
except Exception as e:
    logger.error(f"Falha crítica: {str(e)}")
    processed_sheets = {}