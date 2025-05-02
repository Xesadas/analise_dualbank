import dash
from dash import dcc, html, Input, Output, State, callback, register_page
import dash_bootstrap_components as dbc
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import os
import traceback
import logging
import json
from pathlib import Path


logging.basicConfig(level=logging.DEBUG)
MOUNT_PATH = '/data' if os.environ.get('RENDER') else os.path.join(os.getcwd(), 'data')
EXCEL_PATH = os.path.join(MOUNT_PATH, 'stores.xlsx')

# Função de configuração inicial
def setup_persistent_environment():
    try:
        os.makedirs(MOUNT_PATH, exist_ok=True)
        
        if not os.path.exists(EXCEL_PATH):
            wb = Workbook()
            wb.save(EXCEL_PATH)
        
        if not os.access(MOUNT_PATH, os.W_OK):
            logging.error(f"Sem permissão de escrita em: {MOUNT_PATH}")
            raise PermissionError("Erro de permissão no diretório persistente")

    except Exception as e:
        logging.error(f"Falha na configuração inicial: {str(e)}")
        raise

# Execute a configuração ao iniciar
setup_persistent_environment()

#REFERENTE A ANÁLISE DE DADOS!!!
dash.register_page(
    __name__,
    path='/cadastro',
    title='Cadastro de Clientes',
    name='Cadastro'
)

# Estilos
transaction_style = {
    'border': '1px solid #e0e0e0',
    'borderRadius': '10px',
    'padding': '20px',
    'marginTop': '30px'
}

# =============================================
# LAYOUT
# =============================================
    
layout = dbc.Container([
    html.Div([
        html.Div([
            html.H2("Cadastro de Clientes", className="titulo-dados mb-4"),
            html.Span("Campos obrigatórios*", className="text-muted mb-4 d-block"),
        ], className="container-header text-center mb-5"),
        
        # Seção de Cadastro Principal
        dbc.Card([
            dbc.CardBody([
                html.Div([
                    html.H5("Datas", className="form-section-title"),
                    dbc.Row([
                        dbc.Col([
                            dbc.Label("Data de Cadastro", className="mb-2"),
                            dcc.DatePickerSingle(
                                id='data-cadastro',
                                date=datetime.today(),
                                first_day_of_week=0,
                                display_format='DD/MM/YYYY',
                                className='w-100'
                            )
                        ], md=6, className="mb-4"),
                        
                        dbc.Col([
                            dbc.Label("Data de Aprovação", className="mb-2"),
                            dcc.DatePickerSingle(
                                id='data-aprovacao',
                                first_day_of_week=0,
                                display_format='DD/MM/YYYY',
                                className='w-100'
                            )
                        ], md=6, className="mb-4")
                    ]),
                ]),
                
                html.Div([
                    html.H5("Informações do Estabelecimento", className="form-section-title"),
                    dbc.Row([
                        dbc.Col(
                            dbc.Input(
                                id='nome-estabelecimento',
                                placeholder="Nome do Estabelecimento*",
                                className='mb-4'
                            ), md=12
                        ),
                      
                    ]),
                ]),
                    dbc.Row([
                        dbc.Col(
                            dbc.Input(
                                id='cpf-cnpj',
                                placeholder="CPF/CNPJ*",
                                className='mb-4'
                            ), md=6
                        ),
                        dbc.Col([
                            dbc.Label("Tipo de Comércio*", className="mb-2"),
                            dcc.Dropdown(
                                id='tipo-comercio',
                                options=[
                                    {'label': 'Restaurantes', 'value': 'Restaurantes'},
                                    {'label': 'Cafeterias', 'value': 'Cafeterias'},
                                    {'label': 'Lanchonetes', 'value': 'Lanchonetes'},
                                    {'label': 'Padarias e Confeitarias', 'value': 'Padarias e Confeitarias'},
                                    {'label': 'Mercados e Mercearias', 'value': 'Mercados e Mercearias'},
                                    {'label': 'Lojas de Roupas e Acessórios', 'value': 'Lojas de Roupas e Acessórios'},
                                    {'label': 'Farmácias e Drogarias', 'value': 'Farmácias e Drogarias'},
                                    {'label': 'Salões de Beleza e Barbearias', 'value': 'Salões de Beleza e Barbearias'},
                                    {'label': 'Clínicas Médicas', 'value': 'Clínicas Médicas'},
                                    {'label': 'Pet Shops e Clínicas Veterinárias', 'value': 'Pet Shops e Clínicas Veterinárias'},
                                    {'label': 'Academias e Estúdios de Fitness', 'value': 'Academias e Estúdios de Fitness'},
                                    {'label': 'Lojas de Eletrônicos e Informática', 'value': 'Lojas de Eletrônicos e Informática'},
                                    {'label': 'Oficinas Mecânicas e Auto Peças', 'value': 'Oficinas Mecânicas e Auto Peças'},
                                    {'label': 'Lojas de Materiais de Construção', 'value': 'Lojas de Materiais de Construção'},
                                    {'label': 'Serviços', 'value': 'Serviços'},
                                    {'label': 'Outros', 'value': 'Outros'}
                                ],
                                placeholder="Selecione o Tipo de Comércio*",
                                className='mb-4'
                            )
                        ], md=6)
                    ]),
                
                html.Div([
                    html.H5("Responsável", className="form-section-title"),
                    dbc.Row([
                        dbc.Col(
                            dbc.Input(
                                id='responsavel',
                                placeholder="Responsável do Estabelecimento*",
                                className='mb-4'
                            ), md=6
                        ),
                        dbc.Col(
                            dbc.Input(
                                id='telefone',
                                placeholder="Telefone*",
                                type="tel",
                                className='mb-4'
                            ), md=6
                        )
                    ]),
                    dbc.Row([
                        dbc.Col(
                            dbc.Input(
                                id='cpf-responsavel',
                                placeholder="CPF do Responsável*",
                                className='mb-4'
                            ), md=12
                        )
                    ]),
                ]),
                
                html.Div([
                    html.H5("Representante", className="form-section-title"),
                    dbc.Input(
                        id='representante',
                        placeholder="Nome do Representante*",
                        type="text",
                        className='mb-4'
                    )
                ]),
                
                html.Div([
                    html.H5("Configurações", className="form-section-title"),
                    dbc.Row([
                        dbc.Col([
                            dbc.Label("Portal"),
                            dcc.Dropdown(
                                id='portal',
                                options=[
                                    {'label': 'Ativo', 'value': 'ATIVO'},
                                    {'label': 'Inativo', 'value': 'INATIVO'}
                                ],
                                className='mb-4'
                            )
                        ], md=4),
                        
                        dbc.Col([
                            dbc.Label("PagSeguro"),
                            dcc.Dropdown(
                                id='pagseguro',
                                options=[
                                    {'label': 'Habilitado', 'value': 'HABILITADO'},
                                    {'label': 'Desabilitado', 'value': 'DESABILITADO'}
                                ],
                                className='mb-4'
                            )
                        ], md=4),
                        
                        dbc.Col([
                            dbc.Label("Sub"),
                            dcc.Dropdown(
                                id='sub',
                                options=[
                                    {'label': 'Habilitado', 'value': 'HABILITADO'},
                                    {'label': 'Não Habilitado', 'value': 'NÃO HABILITADO'}
                                ],
                                className='mb-4'
                            )
                        ], md=4)
                    ]),
                    
                    dbc.Row([
                        dbc.Col(
                            dbc.Input(
                                id='pagseguro-email',
                                placeholder="Email PagSeguro",
                                type="email",
                                className='mb-4'
                            ), md=8
                        ),
                        dbc.Col([
                            dbc.Label("Plano PagSeguro"),
                            dcc.Dropdown(
                                id='plano-pagseguro',
                                options=[
                                    {'label': 'NNB', 'value': 'NNB'},
                                    {'label': 'NNA', 'value': 'NNA'},
                                    {'label': 'NNC', 'value': 'NNC'},
                                    {'label': 'NND', 'value': 'NND'}
                                ],
                                className='mb-4'
                            )
                        ], md=4)
                    ]),
                ]),
                
                dbc.Button(
                    "Salvar Cadastro", 
                    id='salvar-button', 
                    className='mt-4',
                    size="lg"
                )
            ])
        ], className="cadastro-container shadow-lg mb-5"),
        
        # Seção de Transações
        dbc.Card([
            dbc.CardBody([
                html.H5("Registro de Transações Diárias", className="form-section-title"),
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Selecionar Cliente"),
                        dcc.Dropdown(
                            id='cliente-transacao',
                            options=[],
                            placeholder="CPF/CNPJ do Cliente*",
                            className='mb-3'
                        )
                    ], md=6),
                    
                    dbc.Col([
                        dbc.Label("Data da Transação"),
                        dcc.DatePickerSingle(
                            id='data-transacao',
                            date=datetime.today(),
                            display_format='DD/MM/YYYY',
                            className='w-100 mb-3'
                        )
                    ], md=3),
                    
                    dbc.Col([
                        dbc.Label("Valor (R$)"),
                        dbc.Input(
                            id='valor-transacao',
                            type='number',
                            step=0.01,
                            placeholder="0.00",
                            className='mb-3'
                        )
                    ], md=3)
                ]),
                dbc.Button(
                    "Registrar Transação",
                    id='salvar-transacao',
                    color="secondary",
                    className='mt-2'
                )
            ])
        ], style=transaction_style, className="shadow-sm mb-4"),
        
        # Seção de Faturamento Mensal
        dbc.Card([
            dbc.CardBody([
                html.H5("Registro de Faturamento Mensal", className="form-section-title"),
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Selecionar Cliente"),
                        dcc.Dropdown(
                            id='cliente-faturamento',
                            options=[],
                            placeholder="CPF/CNPJ do Cliente*",
                            className='mb-3'
                        )
                    ], md=6),
                    
                    dbc.Col([
                        dbc.Label("Mês"),
                        dcc.Dropdown(
                            id='mes-faturamento',
                            options=[
                                {'label': 'Dezembro', 'value': 'Dezembro'},
                                {'label': 'Janeiro', 'value': 'Janeiro'},
                                {'label': 'Fevereiro', 'value': 'Fevereiro'},
                                {'label': 'Março', 'value': 'Marco'},
                                {'label': 'Abril', 'value': 'Abril'},
                                {'label': 'Maio', 'value': 'Maio'},
                                {'label': 'Junho', 'value': 'Junho'},
                                {'label': 'Julho', 'value': 'Julho'},
                                {'label': 'Agosto', 'value': 'Agosto'},
                                {'label': 'Setembro', 'value': 'Setembro'},
                                {'label': 'Outubro', 'value': 'Outubro'},
                                {'label': 'Novembro', 'value': 'Novembro'},
                            ],
                            placeholder="Selecione o Mês*",
                            className='mb-3'
                        )
                    ], md=3),
                    
                    dbc.Col([
                        dbc.Label("Valor (R$)"),
                        dbc.Input(
                            id='valor-faturamento',
                            type='number',
                            step=0.01,
                            placeholder="0.00",
                            className='mb-3'
                        )
                    ], md=3)
                ]),
                dbc.Button(
                    "Salvar Faturamento",
                    id='salvar-faturamento',
                    color="primary",
                    className='mt-2'
                )
            ])
        ], style=transaction_style, className="shadow-sm")
        
    ], className="py-5"),
        
        dbc.Card([
            dbc.CardBody([
                html.H5("Registro de Faturamento Semanal", className="form-section-title"),
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Selecionar Cliente"),
                        dcc.Dropdown(
                            id='cliente-semanal',
                            options=[],
                            placeholder="CPF/CNPJ do Cliente*",
                            className='mb-3'
                        )
                    ], md=5),
                    
                    dbc.Col([
                        dbc.Label("Mês"),
                        dcc.Dropdown(
                            id='mes-semanal',
                            options=[
                                {'label': 'Janeiro', 'value': 'Janeiro'},
                                {'label': 'Fevereiro', 'value': 'Fevereiro'},
                                {'label': 'Março', 'value': 'Marco'},
                                {'label': 'Abril', 'value': 'Abril'},
                                {'label': 'Maio', 'value': 'Maio'},
                                {'label': 'Junho', 'value': 'Junho'},
                                {'label': 'Julho', 'value': 'Julho'},
                                {'label': 'Agosto', 'value': 'Agosto'},
                                {'label': 'Setembro', 'value': 'Setembro'},
                                {'label': 'Outubro', 'value': 'Outubro'},
                                {'label': 'Novembro', 'value': 'Novembro'},
                                {'label': 'Dezembro', 'value': 'Dezembro'},
                            ],
                            placeholder="Selecione o Mês*",
                            className='mb-3'
                        )
                    ], md=3),
                    
                    dbc.Col([
                        dbc.Label("Semana"),
                        dcc.Dropdown(
                            id='semana',
                            options=[{'label': f'Semana {i}', 'value': i} for i in range(1, 6)],
                            placeholder="Nº da Semana*",
                            className='mb-3'
                        )
                    ], md=2),
                    
                    dbc.Col([
                        dbc.Label("Valor (R$)"),
                        dbc.Input(
                            id='valor-semanal',
                            type='number',
                            step=0.01,
                            placeholder="0.00",
                            className='mb-3'
                        )
                    ], md=2)
                ]),
                dbc.Button(
                    "Salvar Semanal",
                    id='salvar-semanal',
                    color="info",
                    className='mt-2'
                )
            ])
        ], style=transaction_style, className="shadow-sm mb-4"),
    
    # Componentes de Armazenamento e Alertas
    dcc.Store(id='clientes-store', storage_type='memory'),
    
    dbc.Alert(
        id='alert', 
        is_open=False, 
        duration=4000, 
        className="animate__animated animate__fadeInRight"
    ),
    dbc.Alert(
        id='alert-transacao', 
        is_open=False, 
        duration=4000,
        className="animate__animated animate__fadeInRight"
    ),
    dbc.Alert(
        id='alert-faturamento', 
        is_open=False, 
        duration=4000,
        className="animate__animated animate__fadeInRight"
    ),
    dbc.Alert(
    id='alert-semanal', 
    is_open=False, 
    duration=4000,
    className="animate__animated animate__fadeInRight"
    )
    
], fluid=True)
# =============================================
# CALLBACKS
# =============================================

@callback(
    Output('cliente-transacao', 'options'),
    Input('clientes-store', 'data')
)
def carregar_clientes(_):
    try:
        if os.path.exists(EXCEL_PATH):  
            df = pd.read_excel(
                EXCEL_PATH,  
                sheet_name='Sheet1',
                usecols=['ESTABELECIMENTO CPF/CNPJ'], 
                dtype={'ESTABELECIMENTO CPF/CNPJ': str}
            )
            
            options = [
                {'label': cnpj, 'value': cnpj} 
                for cnpj in df['ESTABELECIMENTO CPF/CNPJ'].dropna().unique()
                if isinstance(cnpj, str) and cnpj.strip() != ''
            ]
            return options
        return []
    except Exception as e:
        logging.error(f"Erro ao carregar clientes: {str(e)}")
        return []

@callback(
    Output('alert-transacao', 'is_open'),
    Output('alert-transacao', 'children'),
    Output('alert-transacao', 'color'),
    Input('salvar-transacao', 'n_clicks'),
    [
        State('cliente-transacao', 'value'),
        State('data-transacao', 'date'),
        State('valor-transacao', 'value'),
    ],
    prevent_initial_call=True
)
def salvar_transacao(n_clicks, cliente, data_transacao, valor):
    file_path = EXCEL_PATH
    
    if not all([cliente, data_transacao, valor]):
        return True, "Preencha todos os campos obrigatórios! ⚠️", "warning"
    
    try:
        from openpyxl import load_workbook

        # Processar dados
        data_transacao = datetime.strptime(data_transacao.split('T')[0], '%Y-%m-%d').strftime('%d/%m/%Y')
        valor = float(valor)

        # Carregar ou criar arquivo
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            if 'Transacoes' in wb.sheetnames:
                ws = wb['Transacoes']
            else:
                ws = wb.create_sheet('Transacoes')
                ws.append(['CPF/CNPJ', 'DATA', 'VALOR (R$)', 'STATUS'])
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Transacoes'
            ws.append(['CPF/CNPJ', 'DATA', 'VALOR (R$)', 'STATUS'])

        # Adicionar nova transação
        ws.append([cliente, data_transacao, valor, 'PROCESSADO'])

        # Garantir que a planilha principal existe
        if 'Sheet1' not in wb.sheetnames:
            wb.create_sheet('Sheet1')

        # Salvar alterações
        wb.save(file_path)

        return True, f"Transação de R${valor:.2f} registrada com sucesso! ✅", "success"
    
    except Exception as e:
        logging.error(f"Erro: {str(e)}\n{traceback.format_exc()}")
        return True, f"Erro: {str(e)} ❌", "danger"

@callback(
    Output('alert', 'is_open'),
    Output('alert', 'children'),
    Output('alert', 'color'),
    Input('salvar-button', 'n_clicks'),
    [
        State('data-cadastro', 'date'),
        State('data-aprovacao', 'date'),
        State('nome-estabelecimento', 'value'),
        State('cpf-cnpj', 'value'),
        State('tipo-comercio', 'value'),
        State('responsavel', 'value'),
        State('telefone', 'value'),
        State('cpf-responsavel', 'value'),
        State('representante', 'value'),
        State('portal', 'value'),
        State('pagseguro', 'value'),
        State('sub', 'value'),
        State('pagseguro-email', 'value'),
        State('plano-pagseguro', 'value')
    ],
    prevent_initial_call=True
)
def salvar_cadastro(n_clicks, data_cadastro, data_aprovacao, nome_estabelecimento, 
                   cpf_cnpj, tipo_comercio, responsavel, telefone, cpf_responsavel, 
                   representante, portal, pagseguro, sub, pagseguro_email, plano_pagseguro):
    
    file_path = EXCEL_PATH
    
    try:
        from openpyxl import load_workbook

        # Processar datas
        def processar_data(date_str):
            if not date_str:
                return None
            try:
                return datetime.strptime(date_str, '%Y-%m-%d').strftime('%d/%m/%Y')
            except:
                return None

        data_cadastro = processar_data(data_cadastro)
        data_aprovacao = processar_data(data_aprovacao)

        
        novo_registro = {
        'DATA DE CADASTRO': data_cadastro,
        'DATA DE APROVAÇÃO': data_aprovacao,
        'ESTABELECIMENTO NOME1': nome_estabelecimento or '',
        'ESTABELECIMENTO CPF/CNPJ': str(cpf_cnpj).strip() if cpf_cnpj else '',
        'TIPO DE COMÉRCIO': tipo_comercio or 'Outros',  
        'RESPONSÁVEL DO ESTABELECIMENTO': responsavel or '',
        'RESPONSÁVEL E-MAIL': '',  
        'RESPONSÁVEL CPF/CNPJ': cpf_responsavel or '',
        'RESPONSÁVEL TELEFONE': telefone or '',
        'STATUS': 'PENDENTE',
        'REPRESENTANTE NOME1': representante or '',
        'PORTAL': portal or 'INATIVO',
        'PAGSEGURO': pagseguro or 'DESABILITADO',
        'PAGSEGURO EMAIL': pagseguro_email or '',
        'SUB': sub or 'NÃO HABILITADO',
        'BANKING': 'NÃO HABILITADO',  
        'PLANO PAG': plano_pagseguro or '',
        'ATIVIDADE':'', 
        'P S': '',  
        'Faturamento Dezembro': 0,
        'Faturamento Janeiro': 0,
        'Faturamento Fevereiro': 0,
        'Faturamento Março': 0,
        'Faturamento Abril': 0,
        'Faturamento Maio': 0,
        'Faturamento Junho': 0,
        'Faturamento Julho': 0,
        'Faturamento Agosto': 0,
        'Faturamento Setembro': 0,
        'Faturamento Outubro': 0,
        'Faturamento Novembro': 0,

        'Média de Faturamento': 0.0
}

        # Carregar ou criar arquivo
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            if 'Sheet1' in wb.sheetnames:
                ws = wb['Sheet1']
                headers = [cell.value for cell in ws[1]]  # Obter cabeçalhos existentes
            else:
                ws = wb.create_sheet('Sheet1')
                headers = list(novo_registro.keys())
                ws.append(headers)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Sheet1'
            headers = list(novo_registro.keys())
            ws.append(headers)

        # Mapear valores para a ordem dos cabeçalhos
        row_data = []
        for header in headers:
            # Usar valor do novo_registro ou string vazia se não existir
            row_data.append(novo_registro.get(header, ''))  

        ws.append(row_data)

        # Salvar alterações
        wb.save(file_path)

        return True, "Cadastro salvo com sucesso! ✔️", "success"
    
    except Exception as e:
        logging.error(f"Erro: {str(e)}\n{traceback.format_exc()}")
        return True, f"Erro ao salvar: {str(e)} ❌", "danger"
    
@callback(
    Output('cliente-faturamento', 'options'),
    Input('clientes-store', 'data')
)
def carregar_clientes_faturamento(_):
    file_path = EXCEL_PATH
    try:
        if os.path.exists(file_path):
            df = pd.read_excel(
                file_path, 
                sheet_name='Sheet1', 
                usecols=['ESTABELECIMENTO CPF/CNPJ'], 
                dtype={'ESTABELECIMENTO CPF/CNPJ': str}
            )
            
            options = [
                {'label': cnpj, 'value': cnpj} 
                for cnpj in df['ESTABELECIMENTO CPF/CNPJ'].dropna().unique()
                if isinstance(cnpj, str) and cnpj.strip() != ''
            ]
            
            return options
        return []
    except Exception as e:
        logging.error(f"Erro ao carregar clientes (faturamento): {str(e)}")
        return []

@callback(
    Output('alert-faturamento', 'is_open'),
    Output('alert-faturamento', 'children'),
    Output('alert-faturamento', 'color'),
    Input('salvar-faturamento', 'n_clicks'),
    [
        State('cliente-faturamento', 'value'),
        State('mes-faturamento', 'value'),
        State('valor-faturamento', 'value'),
    ],
    prevent_initial_call=True
)
def salvar_faturamento(n_clicks, cliente, mes, valor):
    if not all([cliente, mes, valor is not None]):
        return True, "Preencha todos os campos obrigatórios! ⚠️", "warning"
    
    try:
        from openpyxl import load_workbook

        file_path = EXCEL_PATH
        wb = load_workbook(file_path)
        ws = wb['Sheet1']
        
        # Encontrar coluna do mês
        target_column = f'Faturamento {mes}'
        header = [cell.value for cell in ws[1]]
        try:
            col_idx = header.index(target_column) + 1  # Coluna base 1
        except ValueError:
            wb.close()
            return True, f"Coluna '{target_column}' não existe! ❌", "danger"
        
        # Encontrar linha do cliente
        cpf_cnpj_col = 'ESTABELECIMENTO CPF/CNPJ'
        cpf_cnpj_idx = header.index(cpf_cnpj_col) + 1
        row_found = None
        
        for row in ws.iter_rows(min_row=2):
            current_cpf = str(row[cpf_cnpj_idx - 1].value).strip()
            if current_cpf == cliente.strip():
                row_found = row[0].row
                break
        
        if not row_found:
            wb.close()
            return True, "Cliente não encontrado! ❌", "danger"
        
        # Atualizar célula
        ws.cell(row=row_found, column=col_idx, value=valor)
        wb.save(file_path)
        wb.close()
        
        return True, f"Faturamento de R${valor:.2f} salvo para {mes}! ✅", "success"
    
    except Exception as e:
        logging.error(f"Erro: {str(e)}\n{traceback.format_exc()}")
        return True, f"Erro ao salvar: {str(e)} ❌", "danger"
    
@callback(
    Output('cliente-semanal', 'options'),  # ID do dropdown semanal
    Input('clientes-store', 'data')        # Disparado ao atualizar dados
)
def carregar_clientes_semanal(_):
    file_path = EXCEL_PATH
    
    try:
        if os.path.exists(file_path):
            # Carrega apenas a coluna de CPF/CNPJ
            df = pd.read_excel(
                file_path,
                sheet_name='Sheet1',
                usecols=['ESTABELECIMENTO CPF/CNPJ'],
                dtype={'ESTABELECIMENTO CPF/CNPJ': str}
            )
            
            # Filtra e formata os valores válidos
            options = [
                {'label': cnpj, 'value': cnpj} 
                for cnpj in df['ESTABELECIMENTO CPF/CNPJ'].dropna().unique()
                if isinstance(cnpj, str) and cnpj.strip() != ''
            ]
            
            return options
        
        return []  # Retorna vazio se arquivo não existir
    
    except Exception as e:
        logging.error(f"Erro ao carregar clientes (semanal): {str(e)}")
        return []

    
@callback(
    Output('alert-semanal', 'is_open'),
    Output('alert-semanal', 'children'),
    Output('alert-semanal', 'color'),
    Input('salvar-semanal', 'n_clicks'),
    [
        State('cliente-semanal', 'value'),
        State('mes-semanal', 'value'),
        State('semana', 'value'),
        State('valor-semanal', 'value'),
    ],
    prevent_initial_call=True
)
def salvar_semanal(n_clicks, cliente, mes, semana, valor):
    if not all([cliente, mes, semana, valor is not None]):
        return True, "Preencha todos os campos obrigatórios! ⚠️", "warning"
    
    try:
        from openpyxl import load_workbook
        file_path = EXCEL_PATH
        
        # Nome da aba baseado no mês
        sheet_name = f"Faturamento {mes}"
        
        # Carregar ou criar arquivo
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
                ws.append(['CPF/CNPJ', 'MÊS', 'SEMANA', 'VALOR (R$)', 'DATA REGISTRO'])
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(['CPF/CNPJ', 'MÊS', 'SEMANA', 'VALOR (R$)', 'DATA REGISTRO'])
        
        # Verificar duplicatas
        for row in ws.iter_rows(min_row=2):
            if (str(row[0].value) == cliente and 
                row[2].value == semana and 
                row[1].value == mes):
                wb.close()
                return True, "Já existe registro para esta semana! ⚠️", "warning"
        
        # Adicionar novo registro
        ws.append([
            cliente,
            mes,
            semana,
            float(valor),
            datetime.now().strftime('%d/%m/%Y %H:%M')
        ])
        
        wb.save(file_path)
        wb.close()
        return True, f"Semana {semana} de {mes} salva com R${valor:.2f}! ✅", "success"
    
    except Exception as e:
        logging.error(f"Erro: {str(e)}\n{traceback.format_exc()}")
        return True, f"Erro ao salvar: {str(e)} ❌", "danger"